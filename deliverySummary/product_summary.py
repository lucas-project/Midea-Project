import pandas as pd
import os
from datetime import timedelta
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
import numpy as np
from matplotlib.gridspec import GridSpec
import math
import textwrap
import openpyxl
import re
import datetime
from openpyxl.styles import Font

# Configure matplotlib to use a font that supports Chinese characters
import matplotlib
matplotlib.rcParams['font.sans-serif'] = ['SimSun', 'Arial Unicode MS', 'DejaVu Sans', 'Microsoft YaHei']
matplotlib.rcParams['font.family'] = 'sans-serif'
matplotlib.rcParams['axes.unicode_minus'] = False

# Ignore warnings
import warnings
warnings.filterwarnings('ignore')

def get_week_number(date):
    """Get ISO week number from date"""
    if pd.isna(date):
        return None
    return date.isocalendar()[1]

def get_week_start_end(date):
    """Get start and end date of the week for a given date"""
    if pd.isna(date):
        return None, None
    start = date - timedelta(days=date.weekday())
    end = start + timedelta(days=6)
    return start, end

def get_week_description(start_date, end_date):
    """Generate description for weeks that cross months"""
    if start_date is None or end_date is None:
        return "Unknown Week"
    
    if start_date.month != end_date.month:
        return f"Week {get_week_number(start_date)} ({start_date.strftime('%b %d')} - {end_date.strftime('%b %d')})"
    else:
        return f"Week {get_week_number(start_date)} ({start_date.strftime('%b %d')} - {end_date.strftime('%d')})"

def get_month_from_week_description(week_description):
    """Extract month from week description"""
    if "Unknown Week" in week_description:
        return "Unknown"
    
    # Extract month abbreviation
    try:
        month_abbr = week_description.split('(')[1].split(' ')[0]
        return month_abbr
    except:
        return "Unknown"

def extract_container_num(value):
    """
    Extract ISO container number from mixed 'Contr #' values.
    Improved to handle more container number formats and edge cases.
    """
    if pd.isna(value):
        return ''
    
    s = str(value).strip()
    
    # Handle empty strings
    if not s:
        return ''
    
    # Normalize the string: uppercase and remove extra whitespace
    s = ' '.join(s.upper().split())
    
    # Standard ISO container format: 4 letters followed by 7 digits
    # This is the most common format (e.g., MSCU1234567)
    iso_match = re.search(r'[A-Z]{4}\d{7}', s)
    if iso_match:
        return iso_match.group(0)
    
    # Handle RAC-XXX/YYYY12345 format
    if '/' in s:
        parts = s.split('/', 1)
        container_part = parts[1].strip()
        
        # Check if the part after slash is a valid container number
        iso_match = re.search(r'[A-Z]{4}\d{7}', container_part)
        if iso_match:
            return iso_match.group(0)
        
        # If not a standard format, clean it up and return
        cleaned = re.sub(r'[^A-Z0-9]', '', container_part)
        return cleaned
    
    # Handle cases where container number is just the value itself
    # (e.g., in the Customs sheet 'Container #' column)
    if re.match(r'^[A-Z]{4}\d{7}$', s):
        return s
    
    # For non-standard formats, clean and return
    cleaned = re.sub(r'[^A-Z0-9]', '', s)
    return cleaned

def map_address_to_main_suburb(address):
    """
    Improved function to map VIC warehouse addresses to specific suburbs.
    Handles various inconsistencies found in the address data.
    """
    if pd.isna(address) or not isinstance(address, str) or address.strip() == "" or address == "-":
        return 'Mel (No Specific)'
    
    # Normalize the address: lowercase, remove extra spaces, normalize whitespace
    addr = address.strip().lower()
    addr = re.sub(r'\s+', ' ', addr)
    
    # Special cases
    if any(x in addr for x in ['pending', 'mel warehouse']):
        return 'Mel (No Specific)'
    
    # QLD addresses should not be mapped as VIC
    if 'qld warehouse' in addr:
        return None  # Not a VIC address
    
    # Primary mapping for the three main suburbs by name
    if 'mulgrave' in addr:
        return 'Mulgrave'
    
    if 'dandenong' in addr:
        return 'Dandenong'
    
    if 'laverton' in addr:
        return 'Laverton'
    
    # Check for specific streets/addresses associated with main suburbs
    if any(street in addr for street in ['wellington road', '287-293 wellington']):
        return 'Mulgrave'
    
    if any(street in addr for street in ['ordish road', 'ordish rd', 'arkwright drive', 'arkwright dr']):
        return 'Dandenong'
    
    if any(pattern in addr for pattern in ['gilbertson rd', 'gilbertson road', '1-11 gilbertson', 
                                          'weddel court', 'weddel ct', 'boundary rd', 'boundary road']):
        return 'Laverton'
    
    # Other specific suburbs in VIC
    other_suburbs = {
        'mount waverley': 'Mount Waverley',
        'cranbourne west': 'Cranbourne West',
        'dingley village': 'Dingley Village',
        'truganina': 'Truganina',
        'ravenhall': 'Ravenhall',
        'campbellfield': 'Campbellfield'
    }
    
    for suburb_pattern, suburb_name in other_suburbs.items():
        if suburb_pattern in addr:
            return suburb_name
    
    # Try to extract suburb from VIC postcode format
    vic_pattern = r'(?:,\s*|\s+)([A-Za-z\s]+)(?:\s+VIC\s+\d{4})'
    match = re.search(vic_pattern, addr)
    if match:
        extracted_suburb = match.group(1).strip()
        # Check if the extracted suburb is one of our known suburbs
        for known_suburb in ['mulgrave', 'dandenong', 'laverton', 'mount waverley', 
                            'cranbourne west', 'dingley village', 'truganina', 
                            'ravenhall', 'campbellfield']:
            if known_suburb in extracted_suburb.lower():
                return other_suburbs.get(known_suburb, extracted_suburb.title())
    
    # Default case - no specific suburb identified
    return 'Mel (No Specific)'

def create_product_table_page(product_data, title, page_num, total_pages, pdf):
    """Create a dedicated page for product list as a table"""
    fig = plt.figure(figsize=(11, 8.5))
    
    # Add title
    if page_num == 1:
        plt.suptitle(title, fontsize=16, y=0.98, fontweight='bold')
    else:
        plt.suptitle(f"{title} (Continued)", fontsize=16, y=0.98, fontweight='bold')
    
    # Create table
    ax = plt.subplot(111)
    
    # Create column labels
    col_labels = ['Product Code', 'Quantity', 'Destination']
    
    # Extract table data
    table_data = [[row[0], row[1], row[2]] for row in product_data]
    
    # Create table
    table = ax.table(
        cellText=table_data,
        colLabels=col_labels,
        cellLoc='center',
        loc='center'
    )
    
    # Style the table
    table.auto_set_font_size(False)
    table.set_fontsize(10)
    table.scale(1, 1.5)
    
    # Style the header row
    for j in range(len(col_labels)):
        cell = table[(0, j)]
        cell.set_facecolor('#D3D3D3')
        cell.set_text_props(weight='bold')
    
    # Add alternating row colors for better readability
    for i in range(len(table_data)):
        for j in range(len(col_labels)):
            cell = table[(i+1, j)]
            if i % 2 == 0:
                cell.set_facecolor('#F0F0F0')
    
    # Add page number if multiple pages
    if total_pages > 1:
        plt.figtext(0.5, 0.05, f"Page {page_num} of {total_pages}", ha='center')
    
    ax.axis('off')
    plt.tight_layout(rect=[0, 0, 1, 0.95])
    
    # Save to PDF
    pdf.savefig(fig)
    plt.close()

def get_cross_lined_products(file_path, sheet_name, log_file):
    """Get a list of cross-lined products from the Excel file"""
    crosslined_products = []
    
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        # Check if the sheet exists
        if sheet_name not in wb.sheetnames:
            with open(log_file, 'a', encoding='utf-8') as log:
                log.write(f"Sheet '{sheet_name}' not found in workbook!\n")
            return []
            
        sheet = wb[sheet_name]
        
        # Find the column index for '型号'
        product_col_idx = None
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=1, column=col).value == '型号':
                product_col_idx = col
                break
        
        if product_col_idx is None:
            with open(log_file, 'a', encoding='utf-8') as log:
                log.write(f"Column '型号' not found in sheet {sheet_name}!\n")
            return []
        
        # Check for cross-lined cells in the '型号' column
        for row in range(2, sheet.max_row + 1):  # Start from row 2 (skip header)
            cell = sheet.cell(row=row, column=product_col_idx)
            product_value = cell.value
            
            if product_value is None:
                continue
            
            # Check if cell has strikethrough formatting
            if cell.font and hasattr(cell.font, 'strike') and cell.font.strike:
                crosslined_products.append((sheet_name, product_value))
                with open(log_file, 'a', encoding='utf-8') as log:
                    log.write(f"Found cross-lined product in {sheet_name}: '{product_value}'\n")
        
        return crosslined_products
    
    except Exception as e:
        with open(log_file, 'a', encoding='utf-8') as log:
            log.write(f"Error getting cross-lined products: {str(e)}\n")
        return []

def create_paginated_hierarchical_table(data, columns, title, subtitle, pdf, state_colors):
    """Create a paginated table with a 3-level hierarchy and merged cells."""
    # First create a title page
    fig_title = plt.figure(figsize=(11.69, 8.27))  # A4 landscape
    plt.text(0.5, 0.5, title, fontsize=24, ha='center', va='center', fontweight='bold')
    plt.figtext(0.5, 0.45, subtitle, ha='center', fontsize=12, color='gray')
    plt.axis('off')
    pdf.savefig(fig_title)
    plt.close(fig_title)
    
    # Now create the data pages
    max_rows_per_page = 35
    total_rows = len(data)
    num_pages = (total_rows + max_rows_per_page - 1) // max_rows_per_page

    for page_num in range(num_pages):
        fig, ax = plt.subplots(figsize=(11, 8.5))  # Standard letter size
        ax.axis('off')
        
        # Add margins to create space at top and bottom
        fig.subplots_adjust(left=0.1, right=0.9, top=0.95, bottom=0.05)

        start_row = page_num * max_rows_per_page
        end_row = min(start_row + max_rows_per_page, total_rows)
        page_data = data[start_row:end_row]
        
        if not page_data:
            continue

        # Make columns adaptive with equal width for State and Delivery Destination
        table = ax.table(cellText=page_data, colLabels=columns, cellLoc='center', loc='center', colWidths=[0.15, 0.175, 0.175, 0.35, 0.15])
        table.auto_set_font_size(False)
        table.set_fontsize(8)
        table.scale(0.9, 1.2) # Reduced overall table size

        # --- Row Coloring ---
        # First, identify groups by state
        state_groups = {}
        current_state = None
        start_idx = 0
        
        for i in range(len(page_data)):
            state = page_data[i][0]
            if state != current_state:
                if current_state is not None:
                    state_groups[current_state] = (start_idx, i-1)
                current_state = state
                start_idx = i
                
        # Add the last group
        if current_state is not None:
            state_groups[current_state] = (start_idx, len(page_data)-1)
            
        # Apply colors to entire groups
        for state, (start, end) in state_groups.items():
            row_color = state_colors.get(state, '#FFFFFF')  # Default to white
            for i in range(start, end + 1):
                for j in range(len(columns)):
                    table[i + 1, j].set_facecolor(row_color)


        # --- Merging Logic for 3 levels ---
        if len(page_data) > 1:
            # Level 1 Merge (State)
            level1_groups = []
            current_group_start = 0
            for i in range(1, len(page_data)):
                if page_data[i][0] != page_data[current_group_start][0]:
                    if i - current_group_start > 1:
                        level1_groups.append((current_group_start, i - 1))
                    current_group_start = i
            if len(page_data) - current_group_start > 1:
                level1_groups.append((current_group_start, len(page_data) - 1))

            for start, end in level1_groups:
                table[start + 1, 0].get_text().set_weight('bold') # Make state bold
                for i in range(start + 1, end + 1):
                    table[i + 1, 0].get_text().set_text('')
                table[start + 1, 0].visible_edges = 'LTR'
                for i in range(start + 1, end):
                    table[i + 1, 0].visible_edges = 'LR'
                table[end + 1, 0].visible_edges = 'LBR'

            # Level 2 Merge (Suburb)
            level2_groups = []
            current_group_start = 0
            for i in range(1, len(page_data)):
                # Start new group if state or suburb changes
                if page_data[i][0] != page_data[current_group_start][0] or \
                   page_data[i][1] != page_data[current_group_start][1]:
                    if i - current_group_start > 1:
                        level2_groups.append((current_group_start, i - 1))
                    current_group_start = i
            if len(page_data) - current_group_start > 1:
                level2_groups.append((current_group_start, len(page_data) - 1))
            
            for start, end in level2_groups:
                 table[start + 1, 1].get_text().set_weight('bold') # Make suburb bold
                 for i in range(start + 1, end + 1):
                    table[i + 1, 1].get_text().set_text('')
                 table[start + 1, 1].visible_edges = 'LTR'
                 for i in range(start + 1, end):
                    table[i + 1, 1].visible_edges = 'LR'
                 table[end + 1, 1].visible_edges = 'LBR'

        # Style header
        for j in range(len(columns)):
            table[0, j].set_facecolor('#DDDDDD')
            table[0, j].set_text_props(weight='bold')

        pdf.savefig(fig, bbox_inches='tight')
        plt.close(fig)

def main():
    try:
        current_dir = os.path.dirname(os.path.abspath(__file__))
        input_file = os.path.join(current_dir, '出货汇总表8.25.xlsx')
        output_file = os.path.join(current_dir, 'Container_Product_Summary.pdf')
        log_file = os.path.join(current_dir, 'Container_Product_Summary.txt')
        
        with open(log_file, 'w', encoding='utf-8') as log:
            log.write(f"--- Product Analysis Log ---\n")
            log.write(f"Report generated at: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            
            # --- Load Data ---
            rac_df = pd.read_excel(input_file, sheet_name='2025 RAC')
            mbt_df = pd.read_excel(input_file, sheet_name='2025 MBT')
            customs_df = pd.read_excel(input_file, sheet_name='Customs Decl. & Local Delivery')

            # --- Handle Merged Cells ---
            rac_df['Contr #'] = rac_df['Contr #'].ffill()
            mbt_df['Contr #'] = mbt_df['Contr #'].ffill()

            rac_df.columns = rac_df.columns.str.strip()
            mbt_df.columns = mbt_df.columns.str.strip()
            customs_df.columns = customs_df.columns.str.strip()

            # --- Date Filtering ---
            # Include last month, this month, and all future months
            today = pd.Timestamp.now().normalize()
            first_day_of_current_month = today.replace(day=1)
            first_day_of_last_month = first_day_of_current_month - pd.DateOffset(months=1)
            
            rac_df['ETA'] = pd.to_datetime(rac_df['ETA'], errors='coerce')
            mbt_df['ETA'] = pd.to_datetime(mbt_df['ETA'], errors='coerce')
            
            # Add week number to both dataframes
            rac_df['Week'] = rac_df['ETA'].apply(lambda x: x.isocalendar()[1] if pd.notna(x) else None)
            mbt_df['Week'] = mbt_df['ETA'].apply(lambda x: x.isocalendar()[1] if pd.notna(x) else None)
            
            rac_df_filtered = rac_df[rac_df['ETA'] >= first_day_of_last_month].copy()
            mbt_df_filtered = mbt_df[mbt_df['ETA'] >= first_day_of_last_month].copy()

            # --- Exclude Cross-Lined Products ---
            cross_lined_rac = get_cross_lined_products(input_file, '2025 RAC', log_file)
            cross_lined_mbt = get_cross_lined_products(input_file, '2025 MBT', log_file)
            
            log.write(f"\nFound {len(cross_lined_rac)} cross-lined products in '2025 RAC'.\n")
            log.write(f"Found {len(cross_lined_mbt)} cross-lined products in '2025 MBT'.\n")
            
            rac_df_filtered = rac_df_filtered[~rac_df_filtered['型号'].isin(cross_lined_rac)]
            mbt_df_filtered = mbt_df_filtered[~mbt_df_filtered['型号'].isin(cross_lined_mbt)]

            # --- Suburb Mapping for VIC-Mel ---
            # First clean up the container numbers in all dataframes
            log.write("\n--- Container Number Extraction ---\n")
            
            # Extract container numbers from both RAC and MBT dataframes
            rac_df_filtered['ContainerNum'] = rac_df_filtered['Contr #'].apply(extract_container_num)
            mbt_df_filtered['ContainerNum'] = mbt_df_filtered['Contr #'].apply(extract_container_num)
            
            # Extract container numbers from Customs sheet
            customs_df['ContainerNum'] = customs_df['Container #'].apply(extract_container_num)
            
            # Log some sample container numbers for verification
            log.write("Sample RAC container numbers:\n")
            for i, (orig, extracted) in enumerate(zip(rac_df_filtered['Contr #'].head(5), rac_df_filtered['ContainerNum'].head(5))):
                log.write(f"  {i+1}. Original: {orig} -> Extracted: {extracted}\n")
            
            log.write("\nSample Customs container numbers:\n")
            for i, (orig, extracted) in enumerate(zip(customs_df['Container #'].head(5), customs_df['ContainerNum'].head(5))):
                log.write(f"  {i+1}. Original: {orig} -> Extracted: {extracted}\n")
            
            # Filter customs dataframe for VIC destinations
            # Use exact match for VIC-Mel to avoid issues with other VIC destinations
            customs_df_vic = customs_df[customs_df['Destination'] == 'VIC-Mel'].copy()
            
            # Create container to suburb mapping
            container_to_suburb = {}
            
            # Log mapping information for debugging
            log.write("\n--- Container to Suburb Mapping ---\n")
            log.write(f"Found {len(customs_df_vic)} VIC-Mel entries in Customs sheet\n")
            
            # Create a mapping of container numbers to suburbs
            for _, row in customs_df_vic.iterrows():
                if pd.notna(row['ContainerNum']) and row['ContainerNum'] != '':
                    container_num = row['ContainerNum']
                    suburb = map_address_to_main_suburb(row['Warehouse address'])
                    if suburb:  # Only add if suburb is not None (e.g., QLD addresses)
                        container_to_suburb[container_num] = suburb
                        
                        # Log the mapping for debugging
                        log.write(f"Container: {container_num}, Address: {row['Warehouse address']}, Mapped to: {suburb}\n")
            
            log.write(f"\nCreated mapping for {len(container_to_suburb)} containers\n")
            
            def update_destination(row):
                # Only process VIC-Mel destinations
                if 'VIC-Mel' not in str(row['Destination']):
                    return row['Destination']
                
                # Check if we have a container number
                if pd.isna(row['ContainerNum']) or row['ContainerNum'] == '':
                    log.write(f"No container number for row with destination {row['Destination']}, using 'VIC-Mel (No Specific)'\n")
                    return 'VIC-Mel (No Specific)'
                
                # Look up suburb in our mapping
                container_num = row['ContainerNum']
                suburb = container_to_suburb.get(container_num)
                
                if suburb:
                    result = f"VIC-{suburb}"
                    # Log successful mapping
                    if '型号' in row:
                        product = row['型号'] if pd.notna(row['型号']) else 'Unknown'
                        week = row['Week'] if pd.notna(row['Week']) else 'Unknown'
                        log.write(f"SUCCESS: Product {product} (Week {week}) with container {container_num} mapped to {result}\n")
                    else:
                        log.write(f"Mapped container {container_num} to {result}\n")
                    return result
                else:
                    # Log mapping failure with more details
                    if '型号' in row:
                        product = row['型号'] if pd.notna(row['型号']) else 'Unknown'
                        week = row['Week'] if pd.notna(row['Week']) else 'Unknown'
                        log.write(f"FAILED: Product {product} (Week {week}) with container {container_num} - No suburb mapping found\n")
                    else:
                        log.write(f"No suburb mapping found for container {container_num}, using 'VIC-Mel (No Specific)'\n")
                    
                    # Double-check if this container exists in the Customs sheet at all
                    customs_container = customs_df[customs_df['ContainerNum'] == container_num]
                    if not customs_container.empty:
                        log.write(f"  NOTE: Container {container_num} exists in Customs sheet but with destination: {customs_container['Destination'].iloc[0]}\n")
                    
                    return 'VIC-Mel (No Specific)'

            rac_df_filtered['Destination'] = rac_df_filtered.apply(update_destination, axis=1)
            mbt_df_filtered['Destination'] = mbt_df_filtered.apply(update_destination, axis=1)

            # --- Combine Data ---
            combined_df = pd.concat([rac_df_filtered, mbt_df_filtered], ignore_index=True)
            combined_df = combined_df[['Destination', '型号', '数量', 'Week', 'ETA']].dropna(subset=['Destination', '型号', '数量'])
            
            # --- Aggregate Product Data ---
            # First get the first ETA date for each Week/Destination/Product combination
            eta_df = combined_df.groupby(['Week', 'Destination', '型号'])['ETA'].first().reset_index()
            
            # Then get the sum of quantities
            quantity_df = combined_df.groupby(['Week', 'Destination', '型号'])['数量'].sum().reset_index()
            
            # Merge them back together
            product_summary = pd.merge(quantity_df, eta_df, on=['Week', 'Destination', '型号'], how='left')
            product_summary.rename(columns={'数量': 'Quantity'}, inplace=True)

            # --- Prepare Data for Hierarchical Table ---
            product_summary['State'] = product_summary['Destination'].apply(lambda x: x.split('-')[0] if '-' in x else x)
            product_summary['Suburb'] = product_summary['Destination'].apply(lambda x: x.split('-')[1] if '-' in x else 'N/A')
            
            # Custom sort: put 'Mel (No Specific)' at the end of VIC group
            product_summary['sort_key'] = product_summary['Suburb'].apply(lambda x: 1 if 'No Specific' in x else 0)
            product_summary = product_summary.sort_values(by=['Week', 'State', 'sort_key', 'Suburb', '型号']).drop(columns='sort_key')

            # Format week number for display with date range
            def format_week_with_dates(week_num, eta_date):
                if pd.isna(week_num) or pd.isna(eta_date):
                    return "Unknown Week"
                
                # Get the first day of the week (Monday)
                # Go back to previous Monday if needed
                day_of_week = eta_date.weekday()  # Monday is 0, Sunday is 6
                start_date = eta_date - pd.Timedelta(days=day_of_week)
                end_date = start_date + pd.Timedelta(days=6)  # Sunday
                
                # Format as "Week X (DD/MM - DD/MM)"
                return f"Week {int(week_num)} ({start_date.strftime('%d/%m')} - {end_date.strftime('%d/%m')})"
            
            # Apply the formatting function to create week display with date range
            product_summary['WeekDisplay'] = product_summary.apply(
                lambda row: format_week_with_dates(row['Week'], row['ETA']), axis=1
            )

            # --- Assign colors to states for background styling ---
            unique_states = product_summary['State'].unique()
            color_palette = ['#FFFFFF', '#F0F0F0']  # White and a light grey
            state_colors = {state: color_palette[i % len(color_palette)] for i, state in enumerate(unique_states)}

            table_data = product_summary[['WeekDisplay', 'State', 'Suburb', '型号', 'Quantity']].values.tolist()

            with PdfPages(output_file) as pdf:
                title = "Product Report"
                subtitle = f"{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                columns = ['Week', 'State', 'Delivery Destination', 'Product Code', 'Total Quantity']
                create_paginated_hierarchical_table(table_data, columns, title, subtitle, pdf, state_colors)

            print(f"Product Report generated successfully: {output_file}")
            log.write(f"\nProduct Report generated successfully.\n")

    except FileNotFoundError:
        print(f"Error: The file '{input_file}' was not found.")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
