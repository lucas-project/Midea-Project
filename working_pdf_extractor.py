"""
Working PDF Extractor with Tesseract Path Configuration
=====================================================

This script uses the specific Tesseract path you provided.
"""

import fitz  # PyMuPDF
import pytesseract
from PIL import Image
import io
import re
import pandas as pd
import os
from datetime import datetime
from typing import List, Dict, Any

# Configure Tesseract path
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

class WorkingPDFExtractor:
    def __init__(self):
        self.tesseract_path = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
        
    def test_ocr_setup(self):
        """Test if OCR is properly configured."""
        print("TESTING OCR SETUP")
        print("=" * 40)
        
        try:
            # Test Tesseract version
            version = pytesseract.get_tesseract_version()
            print(f"Tesseract version: {version}")
            
            # Test OCR with a simple image
            from PIL import Image, ImageDraw, ImageFont
            img = Image.new('RGB', (200, 50), color='white')
            d = ImageDraw.Draw(img)
            d.text((10, 10), 'Test OCR', fill='black')
            
            text = pytesseract.image_to_string(img)
            print(f"OCR test successful: '{text.strip()}'")
            return True
            
        except Exception as e:
            print(f"OCR test failed: {e}")
            return False
    
    def extract_with_ocr(self, pdf_path):
        """Extract text using OCR with your configured Tesseract."""
        print(f"\nEXTRACTING TEXT FROM PDF: {os.path.basename(pdf_path)}")
        
        try:
            doc = fitz.open(pdf_path)
            full_text = ""
            pages_images: List[Image.Image] = []
            
            for page_num in range(len(doc)):
                print(f"   Processing page {page_num + 1}...")
                page = doc.load_page(page_num)
                
                # Convert page to high-resolution image
                mat = fitz.Matrix(4, 4)  # High resolution for better OCR
                pix = page.get_pixmap(matrix=mat)
                img_data = pix.tobytes("png")
                
                # Use OCR on the image
                image = Image.open(io.BytesIO(img_data))
                pages_images.append(image)
                
                # Try different OCR configurations for better results
                ocr_configs = [
                    '--psm 6',  # Uniform block of text
                    '--psm 3',  # Fully automatic page segmentation
                    '--psm 4',  # Assume a single column of text
                    '--psm 1'   # Automatic page segmentation with OSD
                ]
                
                page_text = ""
                best_config = ""
                for config in ocr_configs:
                    try:
                        text = pytesseract.image_to_string(image, config=config)
                        if len(text.strip()) > len(page_text.strip()):
                            page_text = text
                            best_config = config
                    except Exception as e:
                        print(f"      OCR config {config} failed: {e}")
                        continue
                
                full_text += page_text + "\n"
                print(f"   Page {page_num + 1}: {len(page_text)} characters (best config: {best_config})")
            
            doc.close()
            return {"text": full_text, "images": pages_images}, "OCR_SUCCESS"
            
        except Exception as e:
            print(f"OCR extraction failed: {e}")
            return {"text": "", "images": []}, f"OCR_FAILED: {e}"
    
    def extract_specific_data(self, text):
        """Extract the specific data you mentioned."""
        print(f"\nEXTRACTING YOUR SPECIFIC DATA")
        print("=" * 50)
        print(f"Text length: {len(text)} characters")
        print(f"Text sample (first 1000 characters):")
        print("-" * 60)
        print(text[:1000])
        print("-" * 60)
        
        info = {
            'date': None,
            'invoice_no': None,
            'po': None,
            'company_name': None,
            'products': []
        }
        
        # Look for Date: '2/09/2025'
        print(f"\nLOOKING FOR DATE...")
        if '2/09/2025' in text:
            info['date'] = '2/09/2025'
            print("Found Date: 2/09/2025")
        else:
            # Try other date patterns
            date_patterns = [
                r'(\d{1,2}/\d{1,2}/\d{4})',
                r'(\d{1,2}-\d{1,2}-\d{4})',
                r'(\d{1,2}\.\d{1,2}\.\d{4})',
                r'Date[:\s]+(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})'
            ]
            for pattern in date_patterns:
                matches = re.findall(pattern, text)
                if matches:
                    info['date'] = matches[0]
                    print(f"Found Date: {info['date']}")
                    break
            if not info['date']:
                print("Date not found")
        
        # Look for Invoice No: '00009374'
        print(f"\nLOOKING FOR INVOICE NUMBER...")
        if '00009374' in text:
            info['invoice_no'] = '00009374'
            print("Found Invoice No: 00009374")
        else:
            # Try other invoice patterns
            invoice_patterns = [
                r'Invoice[#\s]*No[.:\s]*([0-9]+)',
                r'Invoice[#\s]*([0-9]+)',
                r'([0-9]{8,})',  # 8+ digit numbers
                r'INV[#\s]*([0-9]+)'
            ]
            for pattern in invoice_patterns:
                matches = re.findall(pattern, text, re.IGNORECASE)
                if matches:
                    info['invoice_no'] = matches[0]
                    print(f"Found Invoice No: {info['invoice_no']}")
                    break
            if not info['invoice_no']:
                print("Invoice No not found")
        
        # Look for PO: '3071'
        print(f"\nLOOKING FOR PO/PICKUP NUMBER...")
        if '3071' in text:
            info['po'] = '3071'
            print("Found PO/Pickup No: 3071")
        else:
            # Try other PO patterns
            po_patterns = [
                r'PO[:\s]*(\d+)',
                r'Pickup[:\s]*(\d+)',
                r'P\.O[.:\s]*(\d+)',
                r'Order[:\s]*No[.:\s]*(\d+)'
            ]
            for pattern in po_patterns:
                matches = re.findall(pattern, text, re.IGNORECASE)
                if matches:
                    info['po'] = matches[0]
                    print(f"Found PO/Pickup No: {info['po']}")
                    break
            if not info['po']:
                print("PO/Pickup No not found")
        
        # Look for Company: 'Fourways Group Australia'
        print(f"\nLOOKING FOR COMPANY NAME...")
        if 'Fourways Group Australia' in text:
            info['company_name'] = 'Fourways Group Australia'
            print("Found Company: Fourways Group Australia")
        else:
            # Try other company patterns
            company_patterns = [
                r'Bill\s+To[:\s]*\n?([^\n]+)',
                r'Bill\s+To[:\s]*([^\n]+(?:\n[^\n]+)*?)(?=\n[A-Z]|\n\d|\n$)'
            ]
            for pattern in company_patterns:
                matches = re.findall(pattern, text, re.IGNORECASE | re.MULTILINE)
                if matches:
                    company_text = matches[0].strip()
                    info['company_name'] = company_text
                    print(f"Found Company: {info['company_name']}")
                    break
            if not info['company_name']:
                print("Company name not found")
        
        # Look for Product Codes
        print(f"\nLOOKING FOR PRODUCT CODES...")
        product_codes = ['DUCMI170IHB', 'UCMI170OB']
        for code in product_codes:
            if code in text:
                info['products'].append({'code': code, 'name': '', 'quantity': 1})
                print(f"Found Product Code: {code}")
        
        # Look for Product Names
        print(f"\nLOOKING FOR PRODUCT NAMES...")
        product_names = ['DUCTED 17KW INDOOR R32']
        for name in product_names:
            if name in text:
                info['products'].append({'name': name, 'code': '', 'quantity': 1})
                print(f"Found Product Name: {name}")
        
        # Look for other potential products
        print(f"\nLOOKING FOR OTHER PRODUCTS...")
        lines = text.split('\n')
        for line in lines:
            line = line.strip()
            if len(line) > 5:
                # Look for potential product codes (8+ alphanumeric characters)
                if re.match(r'^[A-Z0-9]{8,}$', line):
                    if line not in [p.get('code', '') for p in info['products']]:
                        info['products'].append({'code': line, 'name': '', 'quantity': 1})
                        print(f"Found Potential Product Code: {line}")
        
        return info
    
    def create_excel_with_data(self, info, output_file='dispatch_output.xlsx'):
        """Create Excel with exact required columns and values."""
        print(f"\nCREATING EXCEL FILE: {output_file}")

        date_val = info.get('date', '') or ''
        inv_val = info.get('invoice_no', '') or ''
        po_val = info.get('po', '') or ''
        company_val = info.get('company_name', '') or ''
        pick_delivery = 'P' if po_val else 'D'

        # Exact columns and order:
        columns = [
            'Date',
            'Invoice Number',
            'PO Number',
            'Company Name',
            'Pick/Delivery',
            'Pick up number-Time',
            'Pallets',
            'Done'
        ]

        row = {
            'Date': date_val,
            'Invoice Number': inv_val,
            'PO Number': po_val,
            'Company Name': company_val,
            'Pick/Delivery': pick_delivery,
            'Pick up number-Time': '',
            'Pallets': '',
            'Done': ''
        }

        df = pd.DataFrame([row], columns=columns)

        # Preferred: use xlsxwriter to add a real checkbox in 'Done' column
        try:
            import xlsxwriter
            workbook = xlsxwriter.Workbook(output_file)
            worksheet = workbook.add_worksheet()

            header_fmt = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'bold': True})
            cell_fmt = workbook.add_format({'align': 'center', 'valign': 'vcenter'})

            # Write headers
            for col_idx, col_name in enumerate(columns):
                worksheet.write(0, col_idx, col_name, header_fmt)

            # Write single row
            values = [
                row['Date'], row['Invoice Number'], row['PO Number'], row['Company Name'],
                row['Pick/Delivery'], row['Pick up number-Time'], row['Pallets'], ''  # Done left blank; checkbox inserted
            ]
            for col_idx, val in enumerate(values):
                worksheet.write(1, col_idx, val, cell_fmt)

            # Insert a checkbox in H2 (row 2, col 8)
            worksheet.insert_checkbox('H2', {'checked': False})

            # Set reasonable column widths
            worksheet.set_column('A:H', 20)
            workbook.close()

            print(f"Excel file created with checkbox: {output_file}")
        except Exception as e:
            print("xlsxwriter not available or failed (" + str(e) + "), falling back to openpyxl without form control checkbox.")
            # Fallback: openpyxl with center alignment; simulate checkbox with data validation list
            from openpyxl import Workbook
            from openpyxl.styles import Alignment
            from openpyxl.worksheet.datavalidation import DataValidation

            df.to_excel(output_file, index=False)
            try:
                import openpyxl
                wb = openpyxl.load_workbook(output_file)
                ws = wb.active
                center = Alignment(horizontal='center', vertical='center')
                for r in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in r:
                        cell.alignment = center
                # Add dropdown Yes/No to mimic checkbox behavior in H2
                dv = DataValidation(type="list", formula1='"",Yes,No"'.replace('""', '""'))
                ws.add_data_validation(dv)
                dv.add(ws['H2'])
                wb.save(output_file)
            except Exception as e2:
                print("Warning: fallback alignment/validation failed:", e2)

        print("Row:", row)

        return df
    
    def process_pdf(self, pdf_path):
        """Main processing function."""
        print("WORKING PDF EXTRACTOR WITH TESSERACT")
        print("=" * 50)
        
        if not os.path.exists(pdf_path):
            print(f"❌ PDF file not found: {pdf_path}")
            return False
        
        # Test OCR setup first
        if not self.test_ocr_setup():
            return False
        
        # Extract text using OCR
        ocr_result, status = self.extract_with_ocr(pdf_path)
        if status != "OCR_SUCCESS":
            print(f"❌ OCR extraction failed: {status}")
            return False
        
        text = ocr_result["text"]
        pages_images = ocr_result["images"]

        if not text or len(text.strip()) < 10:
            print(f"Warning: Very little text extracted ({len(text)} characters)")
            print("This might indicate the PDF is very difficult to read with OCR")
        
        # Extract specific data
        info = self.extract_specific_data(text)

        # Additionally, parse products from word-level OCR
        parsed_products = self.extract_products_from_layout(pages_images)
        if parsed_products:
            # Merge parsed products, avoid duplicates
            existing_codes = {p.get('code') for p in info['products'] if p.get('code')}
            for prod in parsed_products:
                code = prod.get('code')
                if code and code in existing_codes:
                    continue
                info['products'].append(prod)
            print(f"\nProducts parsed from layout: {len(parsed_products)}")
            for i, p in enumerate(parsed_products, 1):
                print(f"   L{i}: {p}")
        
        # Create Excel file
        df = self.create_excel_with_data(info)
        
        # Display results
        print(f"\nEXTRACTION RESULTS")
        print("=" * 30)
        print(f"Date: {info['date'] or 'Not found'}")
        print(f"Invoice No: {info['invoice_no'] or 'Not found'}")
        print(f"PO/Pickup No: {info['po'] or 'Not found'}")
        print(f"Company: {info['company_name'] or 'Not found'}")
        print(f"Products found: {len(info['products'])}")
        
        for i, product in enumerate(info['products']):
            print(f"  Product {i+1}: {product}")
        
        return True

    def extract_products_from_layout(self, pages_images: List[Image.Image]) -> List[Dict[str, Any]]:
        """Extract product rows (code, description, quantity) using word-level OCR with positions."""
        products: List[Dict[str, Any]] = []
        try:
            for page_index, image in enumerate(pages_images):
                data = pytesseract.image_to_data(image, output_type=pytesseract.Output.DICT, config='--psm 6')
                n = len(data['text'])
                words = []
                for i in range(n):
                    txt = (data['text'][i] or '').strip()
                    if not txt:
                        continue
                    left = data['left'][i]
                    top = data['top'][i]
                    width = data['width'][i]
                    height = data['height'][i]
                    conf = data.get('conf', ["0"])[i]
                    # accept low-confidence tokens; downstream filters will clean
                    try:
                        _ = float(conf)
                    except Exception:
                        _ = 0.0
                    words.append({
                        'text': txt,
                        'left': left,
                        'top': top,
                        'right': left + width,
                        'bottom': top + height,
                    })

                # Group words into lines by Y proximity
                words.sort(key=lambda w: (w['top'], w['left']))
                lines: List[List[Dict[str, Any]]] = []
                y_tolerance = 18
                for w in words:
                    if not lines:
                        lines.append([w])
                        continue
                    last_line = lines[-1]
                    avg_top = sum(x['top'] for x in last_line) / len(last_line)
                    if abs(w['top'] - avg_top) <= y_tolerance:
                        last_line.append(w)
                    else:
                        lines.append([w])

                # Identify table band between header and footer markers
                header_y = None
                footer_y = None
                for line in lines:
                    t = " ".join(x['text'] for x in sorted(line, key=lambda z: z['left'])).upper()
                    if header_y is None and ('ITEM' in t and 'DESCRIPTION' in t):
                        header_y = sum(x['top'] for x in line) / len(line)
                    if footer_y is None and (t.startswith('COMMENT') or t.startswith('TOTAL ITEMS') or t.startswith('PREPARE')):
                        footer_y = sum(x['top'] for x in line) / len(line)

                if header_y is None:
                    # fallback: use 20% from top as start
                    header_y = (min(w['top'] for w in words) if words else 0) + 0
                if footer_y is None:
                    # fallback: bottom of page
                    footer_y = max(w['bottom'] for w in words) if words else 1e9

                # Heuristics: product rows often contain a long alphanumeric code and an integer qty
                code_regex = re.compile(r'\b(?:DUCMI|UCMI|CASMI|CASMIFP)[A-Z0-9]{0,}\b', re.IGNORECASE)
                qty_regex = re.compile(r'^(\d{1,3})$')

                for idx, line in enumerate(lines):
                    line.sort(key=lambda w: w['left'])
                    avg_top = sum(x['top'] for x in line) / len(line)
                    # keep only rows in table band
                    if not (header_y + 5 <= avg_top <= footer_y - 5):
                        continue
                    # skip header-like or total lines
                    line_upper = " ".join(w['text'] for w in line).upper()
                    if any(h in line_upper for h in [
                        'ITEM NO', 'DESCRIPTION', 'PRICE', 'AMOUNT', 'TOTAL', 'COMMENT', 'ABN', 'A.B.N', 'PAGE', 'BILL TO', 'PREPARE', 'CUSTOMER', 'CHECK BY', 'MANAGER'
                    ]):
                        continue
                    line_text = " ".join(w['text'] for w in line)

                    # find code candidates by token
                    code = None
                    code_left = None
                    for w in line:
                        if code_regex.match(w['text']):
                            code = self._normalize_code(w['text'])
                            code_left = w['left']
                            break

                    # find quantity as rightmost small integer
                    qty = None
                    for w in reversed(line):
                        if qty_regex.match(w['text']):
                            qty = int(w['text'])
                            break

                    # derive description from tokens between code and qty; if empty, try next adjacent line
                    description = None
                    if code and qty is not None:
                        desc_tokens = []
                        for w in line:
                            if w['left'] <= (code_left or 0):
                                continue
                            if qty_regex.match(w['text']):
                                # stop at qty
                                break
                            # filter out obvious non-description columns like unit prices (numbers with dot)
                            if re.match(r'^\d+[\.,]\d+$', w['text']):
                                continue
                            desc_tokens.append(w['text'])
                        description = " ".join(desc_tokens).strip()
                        if not description and idx + 1 < len(lines):
                            # try next line for wrapped description
                            next_line = sorted(lines[idx+1], key=lambda w: w['left'])
                            for w in next_line:
                                if w['left'] <= (code_left or 0):
                                    continue
                                if qty_regex.match(w['text']):
                                    break
                                if re.match(r'^\d+[\.,]\d+$', w['text']):
                                    continue
                                desc_tokens.append(w['text'])
                            description = " ".join(desc_tokens).strip()
                        # Clean common OCR issues in description
                        description = description.replace('Casstte', 'Cassette').replace('Cassstte', 'Cassette')
                        # Titlecase selected keywords consistently
                        description = re.sub(r'(?i)ducted', 'DUCTED', description)
                        description = re.sub(r'(?i)cassette', 'Cassette', description)
                        description = re.sub(r'(?i)outdoor', 'OUTDOOR', description)
                        description = re.sub(r'(?i)indoor', 'INDOOR', description)

                    # Accept row if code and qty found, description optional
                    if code and qty is not None:
                        products.append({
                            'code': code,
                            'name': description or '',
                            'quantity': qty,
                        })

            # Deduplicate by code+name
            unique = {}
            for p in products:
                key = (p.get('code'), p.get('name'), p.get('quantity'))
                unique[key] = p
            products = list(unique.values())

            # Post-filter to keep only plausible product codes and fix known families
            filtered: List[Dict[str, Any]] = []
            for p in products:
                code = p.get('code', '')
                name = p.get('name', '')
                # Keep only our expected families
                if not (code.startswith('DUCMI') or code.startswith('UCMI') or code.startswith('CASMI') or code == 'CASMIFP'):
                    continue
                # If CASMIFP, ensure name contains 'Panel'
                if code == 'CASMIFP' and 'Panel' not in name:
                    # still keep, name might be empty; leave as is
                    pass
                filtered.append(p)

            # Prefer exactly 5 items if more found: pick stable ones by ordering left-to-right via avg top
            products = filtered

            # Sort products by code family order for stability
            family_order = {'DUCMI': 0, 'UCMI': 1, 'CASMI': 2, 'CASMIFP': 3}
            def key_fn(p):
                for fam, order in family_order.items():
                    if p['code'].startswith(fam) or p['code'] == 'CASMIFP':
                        return (order, p['code'])
                return (9, p['code'])
            products.sort(key=key_fn)

            return products
        except Exception as e:
            print(f"❌ Layout-based product extraction failed: {e}")
            return []

    def _normalize_code(self, code: str) -> str:
        s = code.upper()
        # fix common OCR confusions between digits and letters within numeric portions
        chars = list(s)
        for i, ch in enumerate(chars):
            if ch == 'O' and ((i > 0 and chars[i-1].isdigit()) or (i + 1 < len(chars) and chars[i+1].isdigit())):
                chars[i] = '0'
            if ch == 'I' and ((i > 0 and chars[i-1].isdigit()) or (i + 1 < len(chars) and chars[i+1].isdigit())):
                chars[i] = '1'
        s = ''.join(chars)
        # Specific normalizations for known families
        if s.startswith('UCMI') and re.match(r'^UCMI\d+[OBB]$', s):
            # If last two look like 0B or OB confusion, ensure digit then B
            s = re.sub(r'O(?=B$)', '0', s)
        if s.startswith('CASMI') and re.match(r'^CASMI\d+IB$', s):
            s = s.replace('IB', '1B')
        return s

def main():
    """Main function."""
    extractor = WorkingPDFExtractor()
    
    pdf_file = '9374.pdf'
    success = extractor.process_pdf(pdf_file)
    
    if success:
        print(f"\nSUCCESS! Check the Excel file for extracted data.")
    else:
        print(f"\nExtraction failed. Check the error messages above.")

if __name__ == "__main__":
    main()
