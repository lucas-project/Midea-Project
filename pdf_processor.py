"""
Improved PDF Processor with Better Extraction Logic
==================================================

This script focuses on robust extraction logic that can handle various PDF formats
without hardcoding specific product information.
"""

import fitz  # PyMuPDF
import pytesseract
from PIL import Image
import io
import re
import pandas as pd
import os
from datetime import datetime
from typing import List, Dict, Any
import argparse
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Configure Tesseract path
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

class ImprovedPDFProcessor:
    def __init__(self, excel_path="Dispatch Schedule.xlsx"):
        self.excel_path = excel_path
        self.ensure_excel_exists()
    
    def ensure_excel_exists(self):
        """Create the Excel file if it doesn't exist with the required columns."""
        if not os.path.exists(self.excel_path):
            print(f"Creating master Excel file: {self.excel_path}")
            columns = [
                'Date', 'Invoice Number', 'PO Number', 'Company Name',
                'Pick/Delivery', 'Pick up number-Time', 'Pallets', 'Done'
            ]
            df = pd.DataFrame(columns=columns)
            
            writer = pd.ExcelWriter(self.excel_path, engine='openpyxl')
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']
            for cell in worksheet[1]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            writer.close()
            print(f"Master Excel file created: {self.excel_path}")
        else:
            print(f"Using existing Excel file: {self.excel_path}")
    
    def extract_text_from_pdf(self, pdf_path):
        """Extract text from PDF using OCR."""
        try:
            doc = fitz.open(pdf_path)
            full_text = ""
            pages_images = []
            
            for page_num in range(len(doc)):
                page = doc.load_page(page_num)
                
                # Convert page to high-resolution image
                mat = fitz.Matrix(4, 4)
                pix = page.get_pixmap(matrix=mat)
                img_data = pix.tobytes("png")
                
                image = Image.open(io.BytesIO(img_data))
                pages_images.append(image)
                
                # Try different OCR configurations
                ocr_configs = ['--psm 6', '--psm 3', '--psm 4']
                
                page_text = ""
                for config in ocr_configs:
                    try:
                        text = pytesseract.image_to_string(image, config=config)
                        if len(text.strip()) > len(page_text.strip()):
                            page_text = text
                    except Exception as e:
                        print(f"OCR config {config} failed: {e}")
                        continue
                
                full_text += page_text + "\n"
            
            doc.close()
            return full_text, pages_images
            
        except Exception as e:
            print(f"Error extracting text: {e}")
            return "", []
    
    def extract_products_from_text(self, text):
        """
        Extract products from text using improved logic.
        
        This function analyzes the text structure to find product lines
        without hardcoding specific product codes.
        """
        print("\n=== ANALYZING TEXT FOR PRODUCTS ===")
        
        # Split text into lines for analysis
        lines = text.split('\n')
        
        # Find the product table section
        product_section_start = -1
        product_section_end = -1
        
        for i, line in enumerate(lines):
            line_upper = line.strip().upper()
            # Look for table header
            if 'ITEM' in line_upper and 'DESCRIPTION' in line_upper:
                product_section_start = i + 1
                print(f"Found product table header at line {i}: {line.strip()}")
                continue
            
            # Look for end of product section
            if product_section_start > -1 and ('COMMENT' in line_upper or 
                                                'TOTAL ITEMS' in line_upper or 
                                                'PREPARE' in line_upper):
                product_section_end = i
                print(f"Found product table end at line {i}: {line.strip()}")
                break
        
        if product_section_start == -1:
            print("Could not find product table header")
            return []
        
        if product_section_end == -1:
            product_section_end = len(lines)
        
        print(f"Product section: lines {product_section_start} to {product_section_end}")
        
        # Extract product lines from the identified section
        products = []
        product_lines = lines[product_section_start:product_section_end]
        
        print(f"\nAnalyzing {len(product_lines)} lines in product section:")
        for i, line in enumerate(product_lines):
            print(f"  Line {product_section_start + i}: {repr(line)}")
        
        # Pattern to match product lines: quantity | code description
        # This pattern is more flexible and handles various formats
        for i, line in enumerate(product_lines):
            line = line.strip()
            if not line:
                continue
                
            print(f"\nProcessing line: {repr(line)}")
            
            # Look for pattern: number | alphanumeric_code rest_of_line
            # This handles cases where description might be on the same line or next line
            match = re.match(r'^(\d+)\s*\|\s*([A-Z0-9]+)(.*)$', line)
            
            if match:
                quantity = int(match.group(1))
                code = match.group(2).strip()
                description_part = match.group(3).strip()
                
                print(f"  Found product pattern:")
                print(f"    Quantity: {quantity}")
                print(f"    Code: {code}")
                print(f"    Description part: {repr(description_part)}")
                
                # If description is empty or very short, check the next line
                if not description_part or len(description_part) < 3:
                    # Check if next line has more description
                    if i + 1 < len(product_lines):
                        next_line = product_lines[i + 1].strip()
                        print(f"    Checking next line: {repr(next_line)}")
                        
                        # If next line doesn't start with a digit (not another product), 
                        # it might be a continuation of description
                        if next_line and not re.match(r'^\d+\s*\|', next_line):
                            description_part = next_line
                            print(f"    Using next line as description: {repr(description_part)}")
                
                # Clean up description
                description = self.clean_description(description_part)
                
                product = {
                    'code': code,
                    'name': description,
                    'quantity': quantity
                }
                
                products.append(product)
                print(f"  ✓ Added product: {product}")
            else:
                print(f"  No product pattern match")
        
        print(f"\n=== EXTRACTION COMPLETE ===")
        print(f"Found {len(products)} products:")
        for i, product in enumerate(products, 1):
            print(f"  {i}. {product}")
        
        return products
    
    def clean_description(self, description):
        """Clean up product description text."""
        if not description:
            return ""
        
        # Remove extra whitespace
        description = ' '.join(description.split())
        
        # Fix common OCR issues
        description = description.replace('Casstte', 'Cassette')
        description = description.replace('Cassstte', 'Cassette')
        
        # Capitalize common terms consistently
        description = re.sub(r'(?i)\bducted\b', 'DUCTED', description)
        description = re.sub(r'(?i)\bcassette\b', 'Cassette', description)
        description = re.sub(r'(?i)\boutdoor\b', 'OUTDOOR', description)
        description = re.sub(r'(?i)\bindoor\b', 'INDOOR', description)
        description = re.sub(r'(?i)\bpanel\b', 'Panel', description)
        
        return description
    
    def extract_specific_data(self, text):
        """Extract date, invoice, PO, and company information."""
        info = {
            'date': None,
            'invoice_no': None,
            'po': None,
            'company_name': None,
        }
        
        print(f"\n=== EXTRACTING HEADER DATA ===")
        print(f"Text sample (first 500 characters):")
        print("-" * 60)
        print(text[:500])
        print("-" * 60)
        
        # Extract Date
        date_patterns = [
            r'Date[:\s]+(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})',
            r'(\d{1,2}/\d{1,2}/\d{4})',
            r'(\d{1,2}-\d{1,2}-\d{4})',
            r'(\d{1,2}\.\d{1,2}\.\d{4})'
        ]
        
        for pattern in date_patterns:
            matches = re.findall(pattern, text)
            if matches:
                info['date'] = matches[0]
                print(f"✓ Found Date: {info['date']}")
                break
        
        # Extract Invoice Number
        invoice_patterns = [
            r'Invoice[#\s]*No[.:\s]*([0-9]+)',
            r'Invoice[#\s]*([0-9]+)',
            r'([0-9]{8,})',
            r'INV[#\s]*([0-9]+)'
        ]
        
        for pattern in invoice_patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            if matches:
                info['invoice_no'] = matches[0]
                print(f"✓ Found Invoice No: {info['invoice_no']}")
                break
        
        # Extract PO Number
        po_patterns = [
            r'PO[:\s]*(\d+)',
            r'Pickup[:\s]*(\d+)',
            r'P\.O[.:\s]*(\d+)',
            r'Order[:\s]*No[.:\s]*(\d+)'
        ]
        
        for pattern in po_patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            if matches:
                info['po'] = matches[0]
                print(f"✓ Found PO Number: {info['po']}")
                break
        
        # Direct check for specific PO numbers if patterns fail
        if info['po'] is None:
            for line in text.split('\n'):
                line = line.strip()
                if line.isdigit() and len(line) == 4:  # 4-digit PO numbers
                    info['po'] = line
                    print(f"✓ Found PO Number (direct): {info['po']}")
                    break
        
        # Extract Company Name
        company_patterns = [
            r'Bill\s+To[:\s]*\n?([^\n]+)',
            r'Bill\s+To[:\s]*([^\n]+(?:\n[^\n]+)*?)(?=\n[A-Z]|\n\d|\n$)'
        ]
        
        for pattern in company_patterns:
            matches = re.findall(pattern, text, re.IGNORECASE | re.MULTILINE)
            if matches:
                company_text = matches[0].strip()
                info['company_name'] = company_text
                print(f"✓ Found Company: {info['company_name']}")
                break
        
        return info
    
    def process_pdf(self, pdf_path):
        """Process a single PDF and extract information."""
        print(f"\n{'='*60}")
        print(f"PROCESSING PDF: {os.path.basename(pdf_path)}")
        print(f"{'='*60}")
        
        # Extract text from PDF using OCR
        text, images = self.extract_text_from_pdf(pdf_path)
        
        if not text or len(text.strip()) < 10:
            print("⚠️ Warning: Very little text extracted from PDF")
            return None
        
        # Extract header information
        info = self.extract_specific_data(text)
        
        # Extract products
        products = self.extract_products_from_text(text)
        info['products'] = products
        info['product_count'] = sum(p.get('quantity', 0) for p in products)
        
        print(f"\n=== SUMMARY ===")
        print(f"Products found: {len(products)}")
        print(f"Total quantity: {info['product_count']}")
        
        return info
    
    def update_excel(self, info):
        """Update the master Excel file with new information."""
        try:
            date_val = info.get('date', '') or ''
            inv_val = info.get('invoice_no', '') or ''
            po_val = info.get('po', '') or ''
            company_val = info.get('company_name', '') or ''
            
            pick_delivery = 'P' if po_val else 'D'
            print(f"PO Number: {po_val}, Setting Pick/Delivery to: {pick_delivery}")
            
            new_row = {
                'Date': date_val,
                'Invoice Number': inv_val,
                'PO Number': po_val,
                'Company Name': company_val,
                'Pick/Delivery': pick_delivery,
                'Pick up number-Time': '',
                'Pallets': '',
                'Done': ''
            }
            
            # Check for duplicates
            try:
                existing_df = pd.read_excel(self.excel_path)
                if not existing_df.empty and inv_val:
                    if inv_val in existing_df['Invoice Number'].values:
                        print(f"Warning: Invoice {inv_val} already exists. Skipping.")
                        return False
            except Exception:
                pass
            
            # Update Excel with openpyxl for better formatting
            book = load_workbook(self.excel_path)
            sheet = book.active
            
            next_row = sheet.max_row + 1
            columns = ['Date', 'Invoice Number', 'PO Number', 'Company Name', 
                       'Pick/Delivery', 'Pick up number-Time', 'Pallets', 'Done']
            
            for col_idx, col_name in enumerate(columns, 1):
                cell = sheet.cell(row=next_row, column=col_idx)
                cell.value = new_row[col_name]
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            book.save(self.excel_path)
            print(f"✓ Excel updated: Added row for Invoice {inv_val}")
            return True
                
        except Exception as e:
            print(f"Error updating Excel: {e}")
            return False

def main():
    """Main function."""
    parser = argparse.ArgumentParser(description="Process PDF order sheets")
    parser.add_argument("--excel", default="Dispatch Schedule.xlsx", help="Excel file path")
    parser.add_argument("--pdf", help="PDF file to process")
    
    args = parser.parse_args()
    
    processor = ImprovedPDFProcessor(args.excel)
    
    if args.pdf:
        if os.path.exists(args.pdf):
            info = processor.process_pdf(args.pdf)
            if info:
                processor.update_excel(info)
        else:
            print(f"Error: PDF file not found: {args.pdf}")
    else:
        print("Please specify a PDF file with --pdf")

if __name__ == "__main__":
    main()
